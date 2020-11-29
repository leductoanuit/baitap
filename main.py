from logging import root
from re import purge
from openpyxl.workbook import child
from openpyxl.xml.constants import DOC_NS
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import time
import openpyxl
import xml.etree.ElementTree as ET
from pyunpack import Archive

Dong=2
#lay gia tri 1 o trong excel
def get_value_excel(filename, cellname):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['Sheet']
    wb.close()
    return Sheet1[cellname].value
#lay danh sach trang  web
def getListUrl():
    dsURL=[]
    wb = openpyxl.load_workbook('Input_HDDT.xlsx')
    worksheet = wb['Sheet']
    for row in range(2,5):  
          for column in "G":  
               cell_name = "{}{}".format(column, row)
               dsURL.append(worksheet[cell_name].value) 
 
    wb.close()
    return dsURL
#lay danh sach ma tra cuu
def getListCode():
    lsCode=[]
    wb = openpyxl.load_workbook('Input_HDDT.xlsx')
    worksheet = wb['Sheet']
    for row in range(2,5):  
           for column in "D":  
                 cell_name = "{}{}".format(column, row)
                 lsCode.append(worksheet[cell_name].value) 

    wb.close()
    return lsCode

#tai file xml trang web thu nhat
def downLoadXML1(url,code):
    driver = webdriver.Chrome('./chromedriver')
    driver.get(url)
    try:
        myElem = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'IdOfMyElement')))
        print ("Page is ready!")
    except TimeoutException:
      print ("Loading took too much time!")

    inputSearch=driver.find_element_by_id("txtCode")
    inputSearch.send_keys(code)

    btnSearch=driver.find_element_by_id("btnSearchInvoice")
    btnSearch.click()

    driver.find_element_by_xpath("//*[@id='popup-content-container']/div[1]/div[2]/div[9]").click()
    xmlEle=driver.find_element_by_xpath("/html/body/div[12]/div[2]/div/div/div[2]/div[1]/div[2]/div[9]/div/div/div[2]")
    driver.execute_script("arguments[0].click();", xmlEle)
    time.sleep(10)
    driver.close()

#tai xml trang web thu hai 
def downLoadXML2(url,code):
    driver = webdriver.Chrome('./chromedriver')
    driver.get(url)
    try:
       myElem = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'IdOfMyElement')))
       print ("Page is ready!")
    except TimeoutException:
       print ("Loading took too much time!")

    driver.find_element_by_xpath("//*[@id='nguoi_mua_hang']/form/div[3]/div[1]/div/input").send_keys("0001892")
    driver.find_element_by_xpath("//*[@id='nguoi_mua_hang']/form/div[3]/div[2]/div/input").send_keys(code)
    driver.find_element_by_xpath("//*[@id='nguoi_mua_hang']/form/div[5]/div/button").click()
    time.sleep(10)
    js="document.querySelector('#content-wraper > div > div:nth-child(2) > div.container-fluid.padding-content.ng-scope > div.mobile.table_block > div > div:nth-child(3) > div:nth-child(2) > a').click()"
    driver.execute_script(js)
    time.sleep(10)
    driver.close()
 
#tai xml trang web thu ba
def downLoadXML3(url,code):
    driver = webdriver.Chrome('./chromedriver')
    driver.get(url)
    try:
         myElem = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.ID, 'IdOfMyElement')))
         print ("Page is ready!")
    except TimeoutException:
          print ("Loading took too much time!")


    driver.find_element_by_xpath("//*[@id='code']").send_keys(code)
    time.sleep(15)
    driver.find_element_by_xpath("/html/body/div[2]/div[2]/div/div[1]/div/form/div/div/div/div[4]/div[2]/button").click()
    time.sleep(15)
    driver.find_element_by_xpath("/html/body/div[2]/div[2]/div/div[1]/div/div/div[1]/table/tbody/tr/td[8]/a").click()
    time.sleep(10)
    driver.find_element_by_xpath("/html/body/div[2]/div/div/div[2]/button[2]").click()
    time.sleep(10)
    driver.close()

#xuat file excel trang web thu nhat
def exportExcel1():
    global Dong
    lsRow=[]      #mảng chứa nội dung ghi vao excel
    lsItems=[]
       #Lay ten anh
    lsRow.append(get_value_excel('Input_HDDT.xlsx','A2')) #lấy tên ảnh và thêm vào mảng
    tree = ET.parse("../../Downloads/NA18E_0005386.xml") # load file xml vào 
    root=tree.getroot() #get ra root gốc trong xml 
    childrenRoot=root.getchildren()[0] #lấy ra con thứ nhất
    childrenItems=root.getchildren()[0].getchildren()[25] #lấy ra con thứ thứ nhất, trong con thứ nhất đó lấy ra con thứ 25
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}signedDate').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}invoiceNumber').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerLegalName').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerDisplayName').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerTaxCode').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerAddressLine').text)
    wb=openpyxl.load_workbook('Output_tmpl.xlsx')  #mở file excel
    sheet1=wb['Sheet1']  #lấy Sheet1
    listTenSanPham=[]
    dong=Dong
    for i in range(0,len(list(childrenItems))):
           cell_name="{}{}".format("H",dong)
           sheet1[cell_name].value=childrenItems.getchildren()[i].getchildren()[2].text
           dong=dong+1
    
    dong2=Dong
    for i in range(0,len(list(childrenItems))):
           cell_name="{}{}".format("I",dong2)
           sheet1[cell_name].value=childrenItems.getchildren()[i].getchildren()[3].text
           dong2=dong2+1
    
    dong3=Dong
    for i in range(0,len(list(childrenItems))):
           cell_name="{}{}".format("J",dong3)
           sheet1[cell_name].value=childrenItems.getchildren()[i].getchildren()[4].text
           dong3=dong3+1
    
    dong4=Dong
    for i in range(0,len(list(childrenItems))):
           cell_name="{}{}".format("K",dong4)
           sheet1[cell_name].value=childrenItems.getchildren()[i].getchildren()[5].text
           dong4=dong4+1
    
    dong5=Dong
    for i in range(0,len(list(childrenItems))):
           cell_name="{}{}".format("L",dong5)
           sheet1[cell_name].value=childrenItems.getchildren()[i].getchildren()[6].text
           dong5=dong5+1
    
    dong6=Dong
    for i in range(0,len(list(childrenItems))):
           cell_name="{}{}".format("M",dong6)
           sheet1[cell_name].value=childrenItems.getchildren()[i].getchildren()[7].text
           dong6=dong6+1

    #lấy ra nội dung vào thêm vào mảng
    dong7=Dong  
    listCol=['A','B','C','D','E','F','G'] #ds các cột
    for i in range(0,7):
         cell_name="{}{}".format(listCol[i],dong7)  #lấy ra tên ô trong ví dụ khi i=0 thì cell_name là A2,và cứ thế B2,C2,...
         sheet1[cell_name].value=lsRow[i] #gán giá trị cho ô
         dong7=dong7+1
   
    Dong=Dong+len(list(childrenItems))
     
    wb.close()
    wb.save("Output_tmpl.xlsx")

#xuat file excel trang web thu hai
def exportExcel2():
    global Dong
    lsRow=[]
       #Lay ten anh
    lsRow.append(get_value_excel('Input_HDDT.xlsx','A3'))
    tree = ET.parse("../../Downloads/ihoadon.vn_0305162238_0001892_30062019.xml")
    root=tree.getroot()
    childrenRoot=root.getchildren()[1].getchildren()[0].getchildren()[0]
    
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}signedDate').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}invoiceNumber').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerLegalName').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerLegalName').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerTaxCode').text)
    lsRow.append(childrenRoot.find('{http://laphoadon.gdt.gov.vn/2014/09/invoicexml/v1}buyerAddressLine').text)
  
    wb=openpyxl.load_workbook('Output_tmpl.xlsx')
    sheet1=wb['Sheet1']
    listCol=['A','B','C','D','E','F','G']
    for i in range(Dong,len(list(childrenRoot.getchildren()))+Dong):
         k=0
         for j in range(0,7):
             cell_name="{}{}".format(listCol[j],i)
             print(cell_name)
             sheet1[cell_name].value=lsRow[k]
             k=k+1

    
    dong=Dong
    for i in range(0,len(list(childrenRoot.getchildren()[25]))):
             cell_name="{}{}".format("H",dong)
             sheet1[cell_name].value=childrenRoot.getchildren()[25].getchildren()[i].getchildren()[2].text
             dong=dong+1
    
    dong2=Dong
    for i in range(0,len(list(childrenRoot.getchildren()[25]))):
             cell_name="{}{}".format("I",dong2)
             sheet1[cell_name].value=childrenRoot.getchildren()[25].getchildren()[i].getchildren()[3].text
             dong2=dong2+1
    
    dong3=Dong
    for i in range(0,len(list(childrenRoot.getchildren()[25]))):
             cell_name="{}{}".format("J",dong3)
             sheet1[cell_name].value=childrenRoot.getchildren()[25].getchildren()[i].getchildren()[5].text
             dong3=dong3+1

    dong4=Dong
    for i in range(0,len(list(childrenRoot.getchildren()[25]))):
             cell_name="{}{}".format("K",dong4)
             sheet1[cell_name].value=childrenRoot.getchildren()[25].getchildren()[i].getchildren()[6].text
             dong4=dong4+1

    dong5=Dong
    for i in range(0,len(list(childrenRoot.getchildren()[25]))):
             cell_name="{}{}".format("L",dong5)
             sheet1[cell_name].value=childrenRoot.getchildren()[25].getchildren()[i].getchildren()[7].text
             dong5=dong5+1
    

    dong6=Dong
    for i in range(0,len(list(childrenRoot.getchildren()[25]))):
             cell_name="{}{}".format("M",dong6)
             try:
                  sheet1[cell_name].value=childrenRoot.getchildren()[25].getchildren()[i].getchildren()[8].text
             except:
                   sheet1[cell_name]=0

             dong6=dong6+1
    
    Dong=Dong+len(list(childrenRoot.getchildren()[25]))
    wb.close()
    wb.save("Output_tmpl.xlsx")

#xuat file excel trang web thu hai
def exportExcel3():
   lsRow=[]
   Archive('../../Downloads/InHoaDon_25_11.zip').extractall('../../Downloads')
  #Lay ten anh
   lsRow.append(get_value_excel('Input_HDDT.xlsx','A4'))
   tree = ET.parse("../../Downloads/11_2020_01GTKT0-001_TB-19E_24696.xml")
   root=tree.getroot()
   childrenRoot=root.getchildren()[0]
   lsRow.append(childrenRoot.find('SignDate').text)
   lsRow.append(childrenRoot.find('InvoiceNo').text)
   lsRow.append(childrenRoot.find('ComName').text)
   lsRow.append(childrenRoot.find('ComName').text)
   lsRow.append(childrenRoot.find('ComTaxCode').text)
   lsRow.append(childrenRoot.find('ComAddress').text)
   lsRow.append(childrenRoot.getchildren()[34].getchildren()[0].getchildren()[5].text)
   lsRow.append(childrenRoot.getchildren()[34].getchildren()[0].getchildren()[9].text)
   lsRow.append(childrenRoot.getchildren()[34].getchildren()[0].getchildren()[7].text)
   lsRow.append(childrenRoot.getchildren()[34].getchildren()[0].getchildren()[11].text)
   lsRow.append(childrenRoot.getchildren()[34].getchildren()[0].getchildren()[13].text)
   lsRow.append(childrenRoot.getchildren()[34].getchildren()[0].getchildren()[12].text)
   
   wb=openpyxl.load_workbook('Output_tmpl.xlsx')
   sheet1=wb['Sheet1']
   listCol=['A','B','C','D','E','F','G']
   for i in range(0,7):
      cell_name="{}{}".format(listCol[i],Dong)
      sheet1[cell_name].value=lsRow[i]


   dong=Dong
   for i in range(0,len(list(childrenRoot.getchildren()[34]))):
         cell_name="{}{}".format("H",dong)
         sheet1[cell_name].value=childrenRoot.getchildren()[34].getchildren()[i].getchildren()[5].text
         dong=dong+1

   dong2=Dong
   for i in range(0,len(list(childrenRoot.getchildren()[34]))):
         cell_name="{}{}".format("I",dong2)
         sheet1[cell_name].value=childrenRoot.getchildren()[34].getchildren()[i].getchildren()[9].text
         dong2=dong2+1

   dong3=Dong
   for i in range(0,len(list(childrenRoot.getchildren()[34]))):
         cell_name="{}{}".format("J",dong3)
         sheet1[cell_name].value=childrenRoot.getchildren()[34].getchildren()[i].getchildren()[7].text
         dong3=dong3+1


   dong4=Dong
   for i in range(0,len(list(childrenRoot.getchildren()[34]))):
         cell_name="{}{}".format("K",dong4)
         sheet1[cell_name].value=childrenRoot.getchildren()[34].getchildren()[i].getchildren()[11].text
         dong4=dong4+1

   dong5=Dong
   for i in range(0,len(list(childrenRoot.getchildren()[34]))):
         cell_name="{}{}".format("L",dong5)
         sheet1[cell_name].value=childrenRoot.getchildren()[34].getchildren()[i].getchildren()[13].text
         dong5=dong5+1
    
   dong6=Dong
   for i in range(0,len(list(childrenRoot.getchildren()[34]))):
         cell_name="{}{}".format("M",dong6)
         sheet1[cell_name].value=childrenRoot.getchildren()[34].getchildren()[i].getchildren()[12].text
         dong6=dong6+1

   wb.close()
   wb.save("Output_tmpl.xlsx")
   
   
if __name__ == "__main__":
    #listURl=getListUrl()
   # listCode=getListCode()
    #downLoadXML1(listURl[0],listCode[0])
    #downLoadXML2(listURl[1],listCode[1])
    #downLoadXML3(listURl[2],listCode[2])
 exportExcel1()
 # exportExcel2()
  #exportExcel3()
